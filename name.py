import openpyxl
import re
import os
dir=r'C:\Users\86178\Desktop\prog'
name=r'副本-无尽战区重启签名活动——重启 ！让野火不熄.xlsx'
dir=dir.replace('\\','/')   
wb=openpyxl.load_workbook(dir+'/'+name)

print(wb.sheetnames)
num=0
tot=0
mode=re.compile(r'(ID|id)\s*?(:|：| )(.*?)\s')
mode2=re.compile(r'.+(:|：)')
mode3=re.compile(r'(ID|id)\s*?')
mode4=re.compile(r'(.+?)——')
mode5=re.compile(r'[(](.+?)[)]')
ws=wb['签名楼+留言墙+回忆区-展示页']
try:
    f=open(dir+'/ok list.docx','w+')
    f2=open(dir+'/'+'name list.docx','w+')
    for cell in ws['B33':'EI62']:
        for each in cell:
            for cc in str(each.value):
                try:
                    cc.encode(encoding='gbk')
                except UnicodeEncodeError as e:
                    idx=each.value.index(cc)
                    each.value=each.value[:idx]+each.value[idx+1:]
            if(each.value!=None):
                tot+=1
            #print(each.value)
            rt1=mode.search(str(each.value))
            rt2=mode2.search(str(each.value))
            rt3=mode3.search(str(each.value))
            rt4=mode4.search(str(each.value))
            rt5=mode5.search(str(each.value))
            if( rt1!= None):
                print('1 '+each.value)
                #f.write('1 '+each.value+'\r\n')
                #f2.write('1 '+rt1.group(3)+'\r\n')
                #print(rt1.groups())
                num+=1
            elif(rt2 != None):
                print('2 '+each.value)
                #f.write('2 '+each.value+'\r\n')
                num+=1
            elif(rt3 != None):
                print('3 '+each.value)
                #f.write('3 '+each.value+'\r\n')
                num+=1
            elif(rt4 != None):
                print('4 '+each.value)
                #f.write('4 '+each.value+'\r\n')
                num+=1
            elif(rt5 != None):
                print('5 '+each.value)
                #f.write('5 '+each.value+'\r\n')
                num+=1
    print('find %d个\r\ntotal %d'%(num,tot))
finally:
    if f:
        f.close()
    if f2:
        f2.close()
#wb.save(dir+'/'+name)







print('done')





