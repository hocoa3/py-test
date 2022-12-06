import openpyxl
from enum import Enum
import re

#dir2=r'C:\Users\86178\Desktop\'


#dir2=dir2.replace('\\','/')

class Alphabet(Enum):
    A=1
    B=2
    C=3
    D=4
    E=5
    F=6
    G=7
    H=8
    I=9
    J=10
    K=11
    L=12
    M=13
    N=14
    O=15
    P=16
    Q=17
    R=18
    S=19
    T=20
    U=21
    V=22
    W=23
    X=24
    Y=25
    Z=26


def Order_Method_RLC(List):
        for j in range(len(List)):
            for k in range(j+1,len(List)):
                mark1=re.search(r'(\d{1,4}?)-',List[j].name)
                mark2=re.search(r'(\d{1,4}?)-',List[k].name)
                if(mark1==None or mark2==None):
                    if(mark1==None):
                        print(List[j].name)
                    elif(mark1==None):
                        print(List[k].name)
                    break
                else:
                    mark1=mark1.group()
                    mark2=mark2.group()
                    if(mark1>mark2):
                        temp=List[j]
                        List[j]=List[k]
                        List[k]=temp

def Order_Method_VN(List):
    for j in range(len(List)):
            for k in range(j+1,len(List)):
                mark1=re.search(r'N\d{1,2}',List[j].mark)
                mark2=re.search(r'N\d{1,2}',List[k].mark)
                if(mark1==None or mark2==None):
                    if(mark1==None):
                        print(List[j].name)
                    elif(mark1==None):
                        print(List[k].name)
                    break
                else:
                    mark1=mark1.group()[1:]
                    mark2=mark2.group()[1:]
                    if(int(mark1)>int(mark2)):
                        temp=List[j]
                        List[j]=List[k]
                        List[k]=temp

class Component():
    def __init__(self):
        self.mark=None
        self.footprint=None
        self.volume=None
        self.quantity=None
        self.name=None

class BOM_List(Component):
    def __init__(self,filename):
        dir=r'C:\Users\86178\Desktop'
        self.dir=dir.replace('\\','/')
        self.filename=filename
        self._wb=openpyxl.load_workbook(self.dir+'/'+self.filename)
        self._ws=self._wb.active
        self.Component_List=None

    def Read_Component(self):
        start_pos='A6'
        Component_List=[]
        for i in range(6,self._ws.max_row+1):
            New_Component=Component()
            for j in range(4):                                     #D6
                pos=chr(ord(start_pos[0])+j)+str(i)     #j=0是大小,i=j=1是封装，j=2是数量，j=3是下标
                if(j==0):
                    New_Component.volume=self._ws.cell(int(pos[1:]),ord(pos[0])-64).value
                elif(j==1):
                    New_Component.footprint=self._ws.cell(int(pos[1:]),ord(pos[0])-64).value
                elif(j==2):
                    New_Component.quantity=self._ws.cell(int(pos[1:]),ord(pos[0])-64).value
                elif(j==3):
                    New_Component.mark=self._ws.cell(int(pos[1:]),ord(pos[0])-64).value
            Component_List.append(New_Component)
            print(New_Component.mark)
        self.Component_List=Component_List

    def Create_Name(self):
        for x in range(len(self.Component_List)):
            if(re.search('F',str(self.Component_List[x].volume)) and re.match(r'\d{4}',str(self.Component_List[x].footprint))):
                    self.Component_List[x].name='贴片电容'+str(self.Component_List[x].footprint)+'-'+str(self.Component_List[x].volume)
            elif(re.search('F',str(self.Component_List[x].volume)) and re.match(r'\d{3}',str(self.Component_List[x].footprint))):
                    self.Component_List[x].name='贴片电容0'+str(self.Component_List[x].footprint)+'-'+str(self.Component_List[x].volume)
            elif(re.match(r'\d{3} \d{3} \d{3}',str(self.Component_List[x].volume))):
                    self.Component_List[x].name='磁珠0'+str(self.Component_List[x].footprint)+'-'+str(self.Component_List[x].volume)
            elif(re.match(r'\d{4}',str(self.Component_List[x].footprint))):
                    self.Component_List[x].name='贴片电阻'+str(self.Component_List[x].footprint)+'-'+str(self.Component_List[x].volume)+'Ω'
            elif(re.match(r'\d{3}',str(self.Component_List[x].footprint))):
                    self.Component_List[x].name='贴片电阻0'+str(self.Component_List[x].footprint)+'-'+str(self.Component_List[x].volume)+'Ω'
            else:
                    self.Component_List[x].name=str(self.Component_List[x].volume)  
        

    

    def Order_Component(self):                                              #先R后C再L再V(二极管之类)再N(芯片),V也作为其他类型的存储
        R_List=[]
        C_List=[]
        L_List=[]
        V_List=[]
        N_List=[]
        All_List=[R_List,C_List,L_List,V_List,N_List]

        for i in range(len(self.Component_List)):                              #分成几组，在进行内部排序，再把每个小组依次写进去
            if(self.Component_List[i].mark[0]=='R'):
                R_List.append(self.Component_List[i])
            elif(self.Component_List[i].mark[0]=='C'):
                C_List.append(self.Component_List[i])
            elif(self.Component_List[i].mark[0]=='L'):
                L_List.append(self.Component_List[i])
            elif(self.Component_List[i].mark[0]=='V'):
                V_List.append(self.Component_List[i])
            elif(self.Component_List[i].mark[0]=='N'):
                N_List.append(self.Component_List[i])
            else:
                V_List.append(self.Component_List[i])
        Order_Method_RLC(R_List)
        Order_Method_RLC(C_List)
        Order_Method_RLC(L_List)
        Order_Method_VN(V_List)
        Order_Method_VN(N_List)
        for x in range(len(N_List)):
            print(N_List[x].mark)
        return All_List

    def Write_Component(self,List):
        start_pos='B8'
        off=0
        for x in range(5):
            if(x==1):
                leng=len(List[0])
            elif(x==2):
                leng=len(List[0])+len(List[1])
            elif(x==3):
                leng=len(List[0])+len(List[1])+len(List[2])
            elif(x==4):
                leng=len(List[0])+len(List[1])+len(List[2])+len(List[3])
            else:
                leng=0
            for y in range(len(List[x])):
                if(int(List[x][y].quantity) >4):
                    num=List[x][y].quantity
                    for z in range(6):
                        pos=chr(ord(start_pos[0])+z)+str(int(start_pos[1])+y+leng+off)
                        if(z==1):
                            data=List[x][y].mark
                            name=data.split(', ')
                            
                            data=name[0]+', '+name[1]+', '+name[2]+', '+name[3]+', '

                            name=name[4:]
                            print(name[0])
                            num-=4
                            
                        elif(z==0):
                            data=y+leng+5
                        elif(z==3):
                            data=List[x][y].name
                        elif(z==4):
                            data='个'
                        elif(z==5):
                            data=List[x][y].quantity
                        else:
                            data=None
                        self.Write_One_Block(pos,data)
                    off+=1
                    print(num)
                    while(num>4):
                        for z in range(6):
                            pos=chr(ord(start_pos[0])+z)+str(int(start_pos[1])+y+leng+off)
                            if(z==1):
                                data=name[0]+', '+name[1]+', '+name[2]+', '+name[3]+', '
                                name=name[4:]
                                num-=4
                            else:
                                data=None
                            self.Write_One_Block(pos,data)
                        off+=1
                    for z in range(6):
                        pos=chr(ord(start_pos[0])+z)+str(int(start_pos[1])+y+leng+off)
                        if(z==1):
                            if(num==1):
                                data=name[0]
                            else:
                                data=''
                                for m in range(len(name)):
                                    data=data+name[m]+', '
                                data=data[:-2]
                        else:
                            data=None
                        self.Write_One_Block(pos,data)

                    
                else:
                    for z in range(6):
                        pos=chr(ord(start_pos[0])+z)+str(int(start_pos[1])+y+leng+off)
                        if(z==1):
                            data=List[x][y].mark
                        elif(z==0):
                            data=y+leng+5
                        elif(z==3):
                            data=List[x][y].name
                        elif(z==4):
                            data='个'
                        elif(z==5):
                            data=List[x][y].quantity
                        else:
                            data=None
                        self.Write_One_Block(pos,data)
        self.Save()

    def Read_One_Block(self,loc):
        print(self._ws[loc].value)
        return self._ws[loc].value

    def Read_One_Row(self,row):
        List=[]
        for col in range(1,self._ws.max_column+1):
            loc=str(Alphabet(col))[-1]+str(row)
            if(self._ws[loc].value != None):
                print(self._ws[loc].value)
                List.append(self._ws[loc].value)
        return List

    def Read_One_Column(self,col,sta_row=6):
        List=[]
        for row in range(sta_row,self._ws.max_row+1):
            loc=str(Alphabet(col))[-1]+str(row)
            if(self._ws[loc].value != None):
                #print(self._ws[loc].value)
                List.append(self._ws[loc].value)
        return List

    def Write_One_Block(self,pos,data):
        self._ws[pos]=data
        print("%s has rewrited to %s"%(pos,data))
        
   
    def Save(self):
        self._wb.save(self.dir+'/'+self.filename)
        print('save done')

class Template_List(BOM_List):
    def __init__(self):
        self.file=r'C:\Users\86178\Desktop\BOM_Template1.xlsx'
        #self.file=self.file.replace('\\','/')
        
        self._wb=openpyxl.load_workbook(self.file)
        self._ws=self._wb.active

    def Write_OnebyColumn_mark(self,list):
        component_pos='C8'
        pos=component_pos
        off=0
        for x in range(len(list)):
            num=len(list[x].split(', '))
            if(num>=4):
                name=list[x].split(', ')
                
                while(num>4):
                    dat=name[0]+', '+name[1]+', '+name[2]+', '+name[3]+', '
                    self.Write_One_Block(pos,dat)
                    off=off+1
                    pos=pos[0]+str(x+8+off)
                    num-=4
                    name=name[4:]
                if(len(name)==1):
                    dat=name[0]
                else:
                    dat=''
                    for x in range(len(name)):
                        dat=dat+name[x]+', '
                    dat=dat[:-2]
                self.Write_One_Block(pos,dat)
                print('pos1%s'% pos)
            else:
                pos=pos[0]+str(x+8+off)
                self.Write_One_Block(pos,list[x])
                print('pos2%s'% pos)
            pos=pos[0]+str(x+8+1+off)
        print(pos)
        self.Save()

    def Write_OnebyColumn_Quantity(self,list):
        component_pos='G8'
        pos=component_pos
        off=0
        for x in range(len(list)):
            num=list[x]
            self.Write_One_Block(pos,list[x])
            while(num>4):
                pos=pos[0]+str(x+8+1+off)
                off+=1
                self.Write_One_Block(pos,None)
                num-=4
            pos=pos[0]+str(x+8+1+off)
        self.Save()

    def Write_OnebyColumn_number(self,list,quantity):
        component_pos='B4'
        pos=component_pos
        off=0
        idx=0
        num=0
        for x in range(1,len(list)+5):
            if(x>4):
                num=quantity[idx]
            self.Write_One_Block(pos,x)
            while(num>4):
                pos=pos[0]+str(x+4+off)
                off+=1
                self.Write_One_Block(pos,None)
                num-=4
            pos=pos[0]+str(x+4+off)
            if(x>4):
                idx+=1
            print(x)
        self.Save()
    
    def Write_OnebyColumn_unit(self,list,quantity):
        component_pos='F8'
        pos=component_pos
        off=0
        for x in range(0,len(list)):
            num=quantity[x]
            self.Write_One_Block(pos,'个')
            while(num>4):
                pos=pos[0]+str(x+8+1+off)
                off+=1
                self.Write_One_Block(pos,None)
                num-=4
            pos=pos[0]+str(x+8+1+off)
        self.Save()

    def Write_OnebyColumn_VolumeandFootPrint(self,list_V,list_F,quantity):
        component_pos='E8'
        pos=component_pos
        off=0
        for x in range(len(list_V)):
            num=quantity[x]
            if(re.search('F',str(list_V[x])) and re.match(r'\d{4}',str(list_F[x]))):
                name='贴片电容'+str(list_F[x])+'-'+str(list_V[x])
            elif(re.search('F',str(list_V[x])) and re.match(r'\d{3}',str(list_F[x]))):
                name='贴片电容0'+str(list_F[x])+'-'+str(list_V[x])
            elif(re.match(r'\d{3} \d{3} \d{3}',str(list_V[x]))):
                name='磁珠0'+str(list_F[x])+'-'+str(list_V[x])
            elif(re.match(r'\d{4}',str(list_F[x]))):
                name='贴片电阻'+str(list_F[x])+'-'+str(list_V[x])+'Ω'
            elif(re.match(r'\d{3}',str(list_F[x]))):
                name='贴片电阻0'+str(list_F[x])+'-'+str(list_V[x])+'Ω'
            else:
                name=str(list_V[x])  
            self.Write_One_Block(pos,name)
            while(num>4):
                pos=pos[0]+str(x+8+1+off)
                off+=1
                self.Write_One_Block(pos,None)
                num=num -4
            pos=pos[0]+str(x+8+1+off)
        self.Save()

    def Save(self):
        self._wb.save(self.file)
        print('save done')

if __name__=='__main__':
    name1='BOM_Test2.xlsx'
    Template_name=''
    #物料清单源文件名字在此输入
    BOM1=BOM_List(name1)
    
    BOM1.Read_Component()
    BOM1.Create_Name()
    List=BOM1.Order_Component()
    Template=Template_List()
    Template.Write_Component(List)

 
    '''
    mark_list=BOM1.Read_One_Column(4)                                                         #读取下标
    volume_list=BOM1.Read_One_Column(1)                                                       #读取大小
    footprint_list=BOM1.Read_One_Column(2)                                                    #读取封装
    quantity_list=BOM1.Read_One_Column(3)                                                     #读取数量
    Template=Template_List(Template_name)

    Template.Write_OnebyColumn_mark(mark_list)                                                #处理下标
    Template.Write_OnebyColumn_VolumeandFootPrint(volume_list,footprint_list,quantity_list)   #处理封装和大小
    Template.Write_OnebyColumn_Quantity(quantity_list)                                        #处理数量
    Template.Write_OnebyColumn_number(mark_list,quantity_list)                                #处理序号
    Template.Write_OnebyColumn_unit(mark_list,quantity_list)                                  #处理单位
    '''
    print('done')
    
    
    
